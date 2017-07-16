
# openpyxl是python专门用于处理excel的包
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re


# 找出文件夹下所有html后缀的文件
def listFiles(rootdir, prefix='.html'):
    file = []
    if not os.path.exists(rootdir):  # 目录不存在
        return file
    # 获取当前路径中指定后缀的文件
    for path in os.listdir(rootdir):
        """
        os.listdir()——指定所有目录下所有的文件和目录名
        os.path.isfile()——判断指定对象是否为文件。是返回True,否则False
        os.path.join(path, name)——连接目录和文件名。 eg.os.path.join('home/','dex.com')结果为home/dex.com
        """
        if os.path.isfile(os.path.join(rootdir, path)) and path.endswith(prefix):
            file.append(path)
    return file


# 把内容写入到excel中。
def writeExcel(path, content, sheetname='sheet1'):
    # 构造一个workBook的对象
    wb = Workbook()
    sheet = wb.create_sheet(sheetname, 0)  # 构造一个表格
    # 这种写入方式，可以看做是坐标的方式了，记得从1开始的。
    rowCount = 1   # 行下标
    for i in content:
        #print(i)
        colCount = 1  # 列下标
        for j in i:
            #print(j)
            sheet.cell(row=rowCount, column=colCount).value = j  # 写入sheet中
            colCount += 1
        rowCount += 1
    wb.save(path)


# 读取excel的内容
def readExcel(path, sheetname='sheet1'):
    excelcontent = []
    wb2 = load_workbook(path)  # 加载Excel
    ws = wb2.get_sheet_by_name(sheetname)  # 读取sheet
    for i in range(1, ws.max_row + 1):  # 行信息。range的返回是少1个(ws.max_row+1)，还有sheet的坐标是从1开始的。
        rowcontent = []
        for j in range(1, ws.max_column + 1):  # 列信息
            rowcontent.append(ws.cell(row=i, column=j).value)
        excelcontent.append(rowcontent)  # 读取了一整行信息并添加到excelcontent中
    # print("读取数据成功！", excelcontent)
    return excelcontent


# 去除标题中的非法字符 (Windows)
def validateTitle(title):
    rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/\:*?"<>|'
    new_title = re.sub(rstr, "", title)
    return new_title


if __name__ == '__main__':
    print(listFiles(r'D:\workplace\pythonwork\douban_book_catch_zzq\web抓取\文学\小说', 'html')) # 提取该目录下的html文件名
  #   readExcel('D:\workplace\pythonwork\HelloWorld\database/booktag.xlsx')
